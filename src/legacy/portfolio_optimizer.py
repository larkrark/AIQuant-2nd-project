import sys
import logging
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                           QHBoxLayout, QLabel, QLineEdit, QPushButton, 
                           QTextEdit, QComboBox, QProgressBar, QMessageBox,
                           QFileDialog, QGroupBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from datetime import datetime

# 로깅 설정
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('portfolio_optimizer_debug.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def project_weights(weights, lower=None, upper=None):
    """제약조건을 만족하도록 가중치 투영 (종목 수에 따라 동적 조정)"""
    n = len(weights)
    
    # 종목 수에 따른 동적 제약조건
    if lower is None:
        lower = max(0.01, 0.5 / n)  # 최소: 균등배분의 50%
    if upper is None:
        upper = min(0.40, 3.0 / n)  # 최대: 균등배분의 300%
    
    weights = np.clip(weights, lower, upper)
    weights = weights / np.sum(weights)
    return weights


class OptimizationWorker(QThread):
    finished = pyqtSignal(dict)
    progress = pyqtSignal(int, str)
    error = pyqtSignal(str)

    def __init__(self, file_path, selected_model, min_weight=None, max_weight=None):
        super().__init__()
        self.file_path = file_path
        self.selected_model = selected_model
        self.min_weight = min_weight  # None이면 자동 계산
        self.max_weight = max_weight  # None이면 자동 계산

    def run(self):
        try:
            self.progress.emit(10, "엑셀 파일 읽는 중...")
            price_data, returns, tickers, detected_format = self.load_excel_data()
            
            if returns is None or returns.empty:
                self.error.emit("데이터를 읽을 수 없습니다. 파일 형식을 확인해주세요.")
                return
            
            if len(returns) < 30:
                self.error.emit(f"데이터가 부족합니다. (현재: {len(returns)}일, 최소: 30일 필요)")
                return
            
            if len(tickers) < 2:
                self.error.emit(f"종목이 부족합니다. (현재: {len(tickers)}개, 최소: 2개 필요)")
                return
            
            # 비중 제약조건 설정 (None이면 자동 계산)
            n = len(tickers)
            if self.min_weight is None:
                self.min_weight = max(0.01, 0.5 / n)
            if self.max_weight is None:
                self.max_weight = min(0.40, 3.0 / n)
            
            # 제약조건 유효성 검사
            if self.min_weight * n > 1.0:
                self.error.emit(f"최소 비중이 너무 큽니다. {n}종목 × {self.min_weight:.1%} = {self.min_weight*n:.0%} > 100%")
                return
            if self.max_weight * n < 1.0:
                self.error.emit(f"최대 비중이 너무 작습니다. {n}종목 × {self.max_weight:.1%} = {self.max_weight*n:.0%} < 100%")
                return
            
            self.progress.emit(25, f"비중 제약: {self.min_weight:.1%} ~ {self.max_weight:.1%}")
            
            self.progress.emit(20, f"데이터 형식: {detected_format}")
            self.progress.emit(30, "공분산 행렬 계산 중...")
            cov_matrix = returns.cov()
            
            if cov_matrix.isnull().any().any():
                self.error.emit("공분산 행렬 계산 실패. 데이터를 확인해주세요.")
                return
            
            self.progress.emit(50, "포트폴리오 최적화 중...")
            portfolios = {}
            
            if self.selected_model in ['All', 'GMV']:
                self.progress.emit(55, "GMV 최적화 중...")
                weights = self.get_gmv_weights(cov_matrix)
                if weights is not None:
                    portfolios['GMV'] = weights
                    
            if self.selected_model in ['All', 'MDP']:
                self.progress.emit(65, "MDP 최적화 중...")
                weights = self.get_mdp_weights(returns, cov_matrix)
                if weights is not None:
                    portfolios['MDP'] = weights
                    
            if self.selected_model in ['All', 'Risk Parity']:
                self.progress.emit(75, "Risk Parity 최적화 중...")
                weights = self.get_risk_parity_weights(cov_matrix)
                if weights is not None:
                    portfolios['Risk Parity'] = weights
            
            portfolios['Equal-Weight'] = np.ones(len(tickers)) / len(tickers)
            
            if len(portfolios) == 1:
                self.error.emit("최적화 실패. 데이터를 확인해주세요.")
                return
            
            self.progress.emit(90, "성과 평가 중...")
            results = {
                'portfolios': portfolios,
                'returns': returns,
                'cov_matrix': cov_matrix,
                'tickers': tickers,
                'price_data': price_data,
                'detected_format': detected_format,
                'min_weight': self.min_weight,
                'max_weight': self.max_weight
            }
            
            self.progress.emit(100, "최적화 완료!")
            self.finished.emit(results)
            
        except Exception as e:
            logger.error(f"Optimization error: {str(e)}", exc_info=True)
            self.error.emit(str(e))

    def load_excel_data(self):
        """엑셀 파일에서 데이터 로드 (세로형/가로형 자동 감지)"""
        try:
            df = pd.read_excel(self.file_path)
            df.columns = [str(col).strip() for col in df.columns]
            
            # 컬럼명 소문자 버전 생성
            cols_lower = [col.lower() for col in df.columns]
            
            # 형식 자동 감지
            has_date = 'date' in cols_lower
            has_ticker = 'ticker' in cols_lower
            has_price = 'price' in cols_lower
            
            if has_date and has_ticker and has_price:
                # 세로형 (Long format)
                return self._load_vertical_format(df)
            elif has_date or self._is_date_column(df.iloc[:, 0]):
                # 가로형 (Wide format)
                return self._load_horizontal_format(df)
            else:
                # 첫 번째 컬럼이 날짜인지 확인
                if self._is_date_column(df.iloc[:, 0]):
                    return self._load_horizontal_format(df)
                else:
                    raise ValueError(
                        "파일 형식을 인식할 수 없습니다.\n\n"
                        "지원 형식:\n"
                        "1) 세로형: Date, Ticker, Price 컬럼\n"
                        "2) 가로형: Date 컬럼 + 종목코드 컬럼들"
                    )
                    
        except Exception as e:
            logger.error(f"Excel load error: {str(e)}", exc_info=True)
            raise

    def _is_date_column(self, series):
        """컬럼이 날짜 형식인지 확인"""
        try:
            pd.to_datetime(series.head(10))
            return True
        except:
            return False

    def _load_vertical_format(self, df):
        """세로형 데이터 로드 (Date, Ticker, Price)"""
        logger.info("세로형(Long format) 데이터 감지")
        
        # 컬럼명 정규화
        col_map = {}
        for col in df.columns:
            if col.lower() == 'date':
                col_map[col] = 'date'
            elif col.lower() == 'ticker':
                col_map[col] = 'ticker'
            elif col.lower() == 'price':
                col_map[col] = 'price'
        
        df = df.rename(columns=col_map)
        
        df['date'] = pd.to_datetime(df['date'])
        df['price'] = pd.to_numeric(df['price'], errors='coerce')
        df = df.dropna(subset=['date', 'ticker', 'price'])
        df = df.sort_values(['date', 'ticker'])
        
        price_pivot = df.pivot_table(
            index='date', 
            columns='ticker', 
            values='price',
            aggfunc='last'
        )
        
        price_pivot = price_pivot.ffill().bfill().dropna()
        returns = price_pivot.pct_change().dropna()
        returns = returns.replace([np.inf, -np.inf], np.nan).dropna()
        
        tickers = list(returns.columns)
        
        logger.info(f"세로형 로드 완료: {len(returns)}일, {len(tickers)}종목")
        return price_pivot, returns, tickers, "세로형 (Long format)"

    def _load_horizontal_format(self, df):
        """가로형 데이터 로드 (Date + 종목별 컬럼)"""
        logger.info("가로형(Wide format) 데이터 감지")
        
        # 첫 번째 컬럼을 날짜로 사용
        first_col = df.columns[0]
        
        # 날짜 컬럼 찾기
        date_col = None
        if first_col.lower() == 'date':
            date_col = first_col
        elif self._is_date_column(df[first_col]):
            date_col = first_col
        else:
            for col in df.columns:
                if col.lower() == 'date' or self._is_date_column(df[col]):
                    date_col = col
                    break
        
        if date_col is None:
            raise ValueError("날짜 컬럼을 찾을 수 없습니다.")
        
        df[date_col] = pd.to_datetime(df[date_col])
        df = df.set_index(date_col)
        df.index.name = 'date'
        
        # 숫자 컬럼만 선택 (종목 가격)
        price_data = df.select_dtypes(include=[np.number])
        
        if price_data.empty:
            raise ValueError("가격 데이터를 찾을 수 없습니다.")
        
        price_data = price_data.sort_index()
        price_data = price_data.ffill().bfill().dropna()
        
        returns = price_data.pct_change().dropna()
        returns = returns.replace([np.inf, -np.inf], np.nan).dropna()
        
        tickers = list(returns.columns)
        
        logger.info(f"가로형 로드 완료: {len(returns)}일, {len(tickers)}종목")
        return price_data, returns, tickers, "가로형 (Wide format)"

    def get_gmv_weights(self, cov_matrix):
        """GMV 포트폴리오 가중치 계산"""
        try:
            n = len(cov_matrix)
            cov_array = cov_matrix.values
            min_w, max_w = self.min_weight, self.max_weight

            def objective(weights):
                return np.dot(weights.T, np.dot(cov_array, weights))

            def gradient(weights):
                return 2 * np.dot(cov_array, weights)

            def grg_optimize(x0, learning_rate=0.01, max_iter=1000, tol=1e-6):
                weights = x0.copy()
                prev_obj = float('inf')
                for _ in range(max_iter):
                    grad = gradient(weights)
                    weights = weights - learning_rate * grad
                    weights = project_weights(weights, min_w, max_w)
                    curr_obj = objective(weights)
                    if abs(curr_obj - prev_obj) < tol:
                        break
                    prev_obj = curr_obj
                    learning_rate *= 0.999
                return weights, curr_obj

            best_weights, min_variance = None, float('inf')
            for _ in range(30):
                x0 = np.random.uniform(min_w, max_w, n)
                x0 = x0 / np.sum(x0)
                weights, variance = grg_optimize(x0)
                if variance < min_variance:
                    min_variance = variance
                    best_weights = weights
            return best_weights
        except Exception as e:
            logger.error(f"GMV optimization failed: {e}")
            return None

    def get_mdp_weights(self, returns, cov_matrix):
        """MDP 포트폴리오 가중치 계산"""
        try:
            n = len(cov_matrix)
            cov_array = cov_matrix.values
            vols = np.sqrt(np.diag(cov_array))
            min_w, max_w = self.min_weight, self.max_weight

            def objective(weights):
                portfolio_vol = np.sqrt(np.dot(weights.T, np.dot(cov_array, weights)))
                if portfolio_vol < 1e-10:
                    return float('inf')
                return -np.dot(weights, vols) / portfolio_vol

            def gradient(weights):
                port_vol = np.sqrt(np.dot(weights.T, np.dot(cov_array, weights)))
                if port_vol < 1e-10:
                    return np.zeros(n)
                weighted_vols = np.dot(weights, vols)
                grad_port_vol = np.dot(cov_array, weights) / port_vol
                return -(vols * port_vol - weighted_vols * grad_port_vol) / (port_vol ** 2)

            def grg_optimize(x0, learning_rate=0.01, max_iter=1000, tol=1e-6):
                weights = x0.copy()
                prev_obj = float('inf')
                for _ in range(max_iter):
                    grad = gradient(weights)
                    weights = weights - learning_rate * grad
                    weights = project_weights(weights, min_w, max_w)
                    curr_obj = objective(weights)
                    if abs(curr_obj - prev_obj) < tol:
                        break
                    prev_obj = curr_obj
                    learning_rate *= 0.999
                return weights, curr_obj

            best_weights, min_obj = None, float('inf')
            for _ in range(30):
                x0 = np.random.uniform(min_w, max_w, n)
                x0 = x0 / np.sum(x0)
                weights, obj = grg_optimize(x0)
                if obj < min_obj:
                    min_obj = obj
                    best_weights = weights
            return best_weights
        except Exception as e:
            logger.error(f"MDP optimization failed: {e}")
            return None

    def get_risk_parity_weights(self, cov_matrix):
        """Risk Parity 포트폴리오 가중치 계산"""
        try:
            n = len(cov_matrix)
            cov_array = cov_matrix.values
            min_w, max_w = self.min_weight, self.max_weight

            def calc_risk_contribution(weights):
                port_vol = np.sqrt(np.dot(weights.T, np.dot(cov_array, weights)))
                if port_vol < 1e-10:
                    return np.zeros(n)
                marginal_contrib = np.dot(cov_array, weights)
                return np.multiply(marginal_contrib, weights) / port_vol

            def objective(weights):
                risk_contrib = calc_risk_contribution(weights)
                target_risk = 1.0 / n
                return np.sum((risk_contrib - target_risk) ** 2)

            def gradient(weights):
                eps = 1e-8
                grad = np.zeros(n)
                for i in range(n):
                    h = np.zeros(n)
                    h[i] = eps
                    grad[i] = (objective(weights + h) - objective(weights - h)) / (2 * eps)
                return grad

            def grg_optimize(x0, learning_rate=0.01, max_iter=1000, tol=1e-6):
                weights = x0.copy()
                prev_obj = float('inf')
                for _ in range(max_iter):
                    grad = gradient(weights)
                    weights = weights - learning_rate * grad
                    weights = project_weights(weights, min_w, max_w)
                    curr_obj = objective(weights)
                    if abs(curr_obj - prev_obj) < tol:
                        break
                    prev_obj = curr_obj
                    learning_rate *= 0.999
                return weights, curr_obj

            best_weights, min_obj = None, float('inf')
            for _ in range(30):
                x0 = np.random.uniform(min_w, max_w, n)
                x0 = x0 / np.sum(x0)
                weights, obj = grg_optimize(x0)
                if obj < min_obj:
                    min_obj = obj
                    best_weights = weights
            return best_weights
        except Exception as e:
            logger.error(f"Risk Parity optimization failed: {e}")
            return None


class PortfolioOptimizer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Exponential Quant Solutions - 자산배분 최적화 엔진')
        self.setGeometry(100, 100, 1300, 850)
        
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        layout = QHBoxLayout(main_widget)
        
        self.setup_input_panel(layout)
        self.setup_graph_panel(layout)
        
        self.worker = None
        self.status_text = ""
        self.file_path = ""
        self.results = None

    def setup_input_panel(self, layout):
        input_panel = QWidget()
        input_layout = QVBoxLayout(input_panel)
        
        # 파일 선택 그룹
        file_group = QGroupBox("📁 데이터 파일")
        file_layout = QVBoxLayout(file_group)
        
        file_btn_layout = QHBoxLayout()
        self.file_label = QLineEdit()
        self.file_label.setReadOnly(True)
        self.file_label.setPlaceholderText("엑셀 파일을 선택하세요...")
        file_btn_layout.addWidget(self.file_label)
        
        self.browse_button = QPushButton('찾아보기')
        self.browse_button.clicked.connect(self.browse_file)
        file_btn_layout.addWidget(self.browse_button)
        file_layout.addLayout(file_btn_layout)
        
        format_label = QLabel(
            "지원 형식:\n"
            "• 세로형: Date, Ticker, Price 컬럼\n"
            "• 가로형: Date + 종목코드 컬럼들"
        )
        format_label.setStyleSheet("color: gray; font-size: 10px;")
        file_layout.addWidget(format_label)
        input_layout.addWidget(file_group)
        
        # 최적화 설정 그룹
        opt_group = QGroupBox("⚙️ 최적화 설정")
        opt_layout = QVBoxLayout(opt_group)
        opt_layout.addWidget(QLabel('Optimization Models:'))
        self.model_combo = QComboBox()
        self.model_combo.addItems(['All', 'GMV', 'MDP', 'Risk Parity'])
        opt_layout.addWidget(self.model_combo)
        input_layout.addWidget(opt_group)
        
        # 비중 제약 설정 그룹
        constraint_group = QGroupBox("📊 비중 제약조건")
        constraint_layout = QVBoxLayout(constraint_group)
        
        # 자동/수동 선택
        self.auto_constraint_check = QPushButton("✅ 자동 설정 (종목 수 기반)")
        self.auto_constraint_check.setCheckable(True)
        self.auto_constraint_check.setChecked(True)
        self.auto_constraint_check.clicked.connect(self.toggle_constraint_mode)
        self.auto_constraint_check.setStyleSheet("""
            QPushButton { text-align: left; padding: 5px; border: 1px solid #ccc; border-radius: 3px; }
            QPushButton:checked { background-color: #e8f5e9; border-color: #4CAF50; }
        """)
        constraint_layout.addWidget(self.auto_constraint_check)
        
        # 수동 입력 필드
        manual_layout = QHBoxLayout()
        manual_layout.addWidget(QLabel('최소:'))
        self.min_weight_input = QLineEdit()
        self.min_weight_input.setText("5")  # 기본값
        self.min_weight_input.setEnabled(False)
        self.min_weight_input.setFixedWidth(50)
        manual_layout.addWidget(self.min_weight_input)
        manual_layout.addWidget(QLabel('%'))
        
        manual_layout.addSpacing(10)
        manual_layout.addWidget(QLabel('최대:'))
        self.max_weight_input = QLineEdit()
        self.max_weight_input.setText("40")  # 기본값
        self.max_weight_input.setEnabled(False)
        self.max_weight_input.setFixedWidth(50)
        manual_layout.addWidget(self.max_weight_input)
        manual_layout.addWidget(QLabel('%'))
        manual_layout.addStretch()
        
        constraint_layout.addLayout(manual_layout)
        
        # 안내 문구 (자동 계산값 표시용)
        self.constraint_hint = QLabel("💡 파일 선택 후 종목 수에 따라 자동 계산됩니다")
        self.constraint_hint.setStyleSheet("color: #666; font-size: 10px;")
        constraint_layout.addWidget(self.constraint_hint)
        
        input_layout.addWidget(constraint_group)
        
        # 버튼 그룹
        btn_layout = QHBoxLayout()
        
        self.run_button = QPushButton('🚀 Run')
        self.run_button.setStyleSheet("""
            QPushButton { background-color: #4CAF50; color: white; font-weight: bold; padding: 10px; border-radius: 5px; }
            QPushButton:hover { background-color: #45a049; }
            QPushButton:disabled { background-color: #cccccc; }
        """)
        self.run_button.clicked.connect(self.start_optimization)
        btn_layout.addWidget(self.run_button)
        
        self.export_button = QPushButton('📥 Export Excel')
        self.export_button.setStyleSheet("""
            QPushButton { background-color: #2196F3; color: white; font-weight: bold; padding: 10px; border-radius: 5px; }
            QPushButton:hover { background-color: #1976D2; }
            QPushButton:disabled { background-color: #cccccc; }
        """)
        self.export_button.clicked.connect(self.export_results)
        self.export_button.setEnabled(False)
        btn_layout.addWidget(self.export_button)
        
        input_layout.addLayout(btn_layout)
        
        # 프로그레스 바
        self.progress_bar = QProgressBar()
        input_layout.addWidget(self.progress_bar)
        
        # 결과 텍스트
        result_group = QGroupBox("📊 결과")
        result_layout = QVBoxLayout(result_group)
        self.result_text = QTextEdit()
        self.result_text.setReadOnly(True)
        self.result_text.setStyleSheet("font-family: Consolas, monospace;")
        result_layout.addWidget(self.result_text)
        input_layout.addWidget(result_group)
        
        layout.addWidget(input_panel, 1)

    def setup_graph_panel(self, layout):
        self.figure, self.ax = plt.subplots(2, 2, figsize=(10, 8))
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas, 2)

    def browse_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "엑셀 파일 선택", "",
            "Excel Files (*.xlsx *.xls);;All Files (*)"
        )
        if file_path:
            self.file_path = file_path
            self.file_label.setText(file_path)
            self.update_status(f"파일 선택됨: {file_path}")
            
            # 종목 수 미리 파악하여 자동 제약조건 표시
            self.preview_constraints()

    def preview_constraints(self):
        """파일에서 종목 수를 파악하여 자동 제약조건 미리보기"""
        try:
            df = pd.read_excel(self.file_path)
            df.columns = [str(col).strip().lower() for col in df.columns]
            
            # 종목 수 파악
            if 'ticker' in df.columns:
                # 세로형
                n_tickers = df['ticker'].nunique()
            else:
                # 가로형 (첫 번째 컬럼이 날짜, 나머지가 종목)
                n_tickers = len(df.columns) - 1
            
            if n_tickers > 0:
                # 자동 계산값
                auto_min = max(0.01, 0.5 / n_tickers) * 100
                auto_max = min(0.40, 3.0 / n_tickers) * 100
                
                self.constraint_hint.setText(
                    f"💡 {n_tickers}종목 감지 → 자동 설정 시: {auto_min:.1f}% ~ {auto_max:.1f}%"
                )
                self.constraint_hint.setStyleSheet("color: #2196F3; font-size: 10px; font-weight: bold;")
                
                # 자동 모드면 입력 필드에도 표시 (읽기 전용)
                if self.auto_constraint_check.isChecked():
                    self.min_weight_input.setText(f"{auto_min:.1f}")
                    self.max_weight_input.setText(f"{auto_max:.1f}")
                    
        except Exception as e:
            self.constraint_hint.setText(f"💡 파일 미리보기 실패: {str(e)[:30]}")
            self.constraint_hint.setStyleSheet("color: orange; font-size: 10px;")

    def toggle_constraint_mode(self):
        """자동/수동 모드 전환"""
        is_auto = self.auto_constraint_check.isChecked()
        self.min_weight_input.setEnabled(not is_auto)
        self.max_weight_input.setEnabled(not is_auto)
        
        if is_auto:
            self.auto_constraint_check.setText("✅ 자동 설정 (종목 수 기반)")
            # 파일이 선택되어 있으면 자동 값으로 업데이트
            if self.file_path:
                self.preview_constraints()
            else:
                self.constraint_hint.setText("💡 파일 선택 후 종목 수에 따라 자동 계산됩니다")
                self.constraint_hint.setStyleSheet("color: #666; font-size: 10px;")
        else:
            self.auto_constraint_check.setText("⬜ 수동 설정")
            self.constraint_hint.setText("💡 수동: 원하는 최소/최대 비중(%)을 직접 입력하세요")
            self.constraint_hint.setStyleSheet("color: #666; font-size: 10px;")

    def start_optimization(self):
        if not self.file_path:
            self.show_error("엑셀 파일을 먼저 선택해주세요.")
            return
        
        # 비중 제약 파싱
        min_weight = None
        max_weight = None
        
        if self.auto_constraint_check.isChecked():
            # 자동 모드: None 전달 → Worker에서 자동 계산
            min_weight = None
            max_weight = None
            constraint_msg = "자동 설정"
        else:
            # 수동 모드: 입력값 사용
            try:
                min_text = self.min_weight_input.text().strip()
                max_text = self.max_weight_input.text().strip()
                
                if not min_text or not max_text:
                    self.show_error("수동 모드에서는 최소/최대 비중을 모두 입력해주세요.")
                    return
                
                min_weight = float(min_text) / 100.0
                max_weight = float(max_text) / 100.0
                
                if min_weight < 0 or min_weight > 1:
                    self.show_error("최소 비중은 0~100% 사이여야 합니다.")
                    return
                
                if max_weight < 0 or max_weight > 1:
                    self.show_error("최대 비중은 0~100% 사이여야 합니다.")
                    return
                
                if min_weight >= max_weight:
                    self.show_error("최소 비중은 최대 비중보다 작아야 합니다.")
                    return
                
                constraint_msg = f"수동 설정 ({min_text}% ~ {max_text}%)"
                    
            except ValueError:
                self.show_error("비중은 숫자로 입력해주세요. (예: 5)")
                return
        
        self.run_button.setEnabled(False)
        self.export_button.setEnabled(False)
        self.progress_bar.setValue(0)
        self.status_text = ""
        self.update_status(f"최적화 시작...")
        self.update_status(f"비중 제약: {constraint_msg}")
        
        self.worker = OptimizationWorker(
            self.file_path, 
            self.model_combo.currentText(),
            min_weight,
            max_weight
        )
        self.worker.progress.connect(self.update_progress)
        self.worker.finished.connect(self.optimization_complete)
        self.worker.error.connect(self.show_error)
        self.worker.start()

    def update_status(self, message):
        self.status_text += message + "\n"
        self.result_text.setPlainText(self.status_text)
        self.result_text.verticalScrollBar().setValue(
            self.result_text.verticalScrollBar().maximum()
        )

    def update_progress(self, value, message):
        self.progress_bar.setValue(value)
        self.update_status(message)

    def show_error(self, error_message):
        self.update_status(f"❌ 오류: {error_message}")
        QMessageBox.critical(self, "Error", error_message)
        self.run_button.setEnabled(True)
        self.progress_bar.setValue(0)

    def optimization_complete(self, results):
        self.results = results
        self.run_button.setEnabled(True)
        self.export_button.setEnabled(True)
        self.update_results(results)
        self.update_plots(results)

    def update_results(self, results):
        portfolios = results['portfolios']
        returns = results['returns']
        tickers = results['tickers']
        price_data = results['price_data']
        
        output = ["\n" + "="*50, "         최적화 결과", "="*50]
        output.append(f"\n📊 데이터 정보:")
        output.append(f"   형식: {results['detected_format']}")
        output.append(f"   기간: {price_data.index[0].strftime('%Y-%m-%d')} ~ {price_data.index[-1].strftime('%Y-%m-%d')}")
        output.append(f"   거래일: {len(returns)}일")
        output.append(f"   종목: {', '.join(tickers)}")
        output.append(f"   비중 제약: {results['min_weight']:.1%} ~ {results['max_weight']:.1%}")
        
        for name, weights in portfolios.items():
            if weights is None:
                continue
            
            output.append(f"\n{'─'*50}")
            output.append(f"📈 {name} Portfolio")
            output.append(f"{'─'*50}")
            output.append("\n  [가중치]")
            
            for ticker, weight in zip(tickers, weights):
                bar = "█" * int(weight * 30)
                output.append(f"  {ticker:8s}: {weight:6.2%} {bar}")
            
            port_returns = returns.dot(weights)
            cum_ret = (1 + port_returns).cumprod()
            total_ret = cum_ret.iloc[-1] - 1
            n_days = len(returns)
            ann_ret = (1 + total_ret) ** (252 / n_days) - 1
            ann_vol = port_returns.std() * np.sqrt(252)
            sharpe = ann_ret / ann_vol if ann_vol > 0 else 0
            mdd = ((cum_ret - cum_ret.cummax()) / cum_ret.cummax()).min()
            
            output.append(f"\n  [성과 지표]")
            output.append(f"  Annual Return:    {ann_ret:>8.2%}")
            output.append(f"  Annual Volatility:{ann_vol:>8.2%}")
            output.append(f"  Sharpe Ratio:     {sharpe:>8.2f}")
            output.append(f"  Max Drawdown:     {mdd:>8.2%}")
        
        output.append("\n" + "="*50)
        self.status_text += "\n".join(output)
        self.result_text.setPlainText(self.status_text)
        self.result_text.verticalScrollBar().setValue(
            self.result_text.verticalScrollBar().maximum()
        )

    def export_results(self):
        """결과를 엑셀 파일로 내보내기"""
        if self.results is None:
            self.show_error("먼저 최적화를 실행해주세요.")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "결과 저장", 
            f"portfolio_result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            self._create_result_excel(file_path)
            self.update_status(f"\n✅ 결과 저장 완료: {file_path}")
            QMessageBox.information(self, "완료", f"결과가 저장되었습니다.\n{file_path}")
        except Exception as e:
            self.show_error(f"저장 실패: {str(e)}")

    def _create_result_excel(self, file_path):
        """결과 엑셀 파일 생성"""
        wb = Workbook()
        
        # 스타일 정의
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4472C4')
        subheader_fill = PatternFill('solid', fgColor='8EA9DB')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        pct_format = '0.00%'
        num_format = '0.00'
        
        results = self.results
        portfolios = results['portfolios']
        returns = results['returns']
        tickers = results['tickers']
        price_data = results['price_data']
        
        # ========== Sheet 1: 비중 요약 ==========
        ws1 = wb.active
        ws1.title = "비중 요약"
        
        # 헤더
        headers = ['종목'] + list(portfolios.keys())
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        # 데이터
        for row, ticker in enumerate(tickers, 2):
            ws1.cell(row=row, column=1, value=ticker).border = border
            for col, (name, weights) in enumerate(portfolios.items(), 2):
                if weights is not None:
                    cell = ws1.cell(row=row, column=col, value=weights[row-2])
                    cell.number_format = pct_format
                    cell.border = border
                    cell.alignment = Alignment(horizontal='right')
        
        # 합계 행
        sum_row = len(tickers) + 2
        ws1.cell(row=sum_row, column=1, value='합계').font = Font(bold=True)
        ws1.cell(row=sum_row, column=1).border = border
        for col in range(2, len(portfolios) + 2):
            col_letter = get_column_letter(col)
            cell = ws1.cell(row=sum_row, column=col, 
                           value=f'=SUM({col_letter}2:{col_letter}{sum_row-1})')
            cell.number_format = pct_format
            cell.font = Font(bold=True)
            cell.border = border
        
        # 열 너비
        ws1.column_dimensions['A'].width = 12
        for col in range(2, len(portfolios) + 2):
            ws1.column_dimensions[get_column_letter(col)].width = 15
        
        # ========== Sheet 2: 성과 지표 ==========
        ws2 = wb.create_sheet("성과 지표")
        
        metrics_headers = ['지표'] + list(portfolios.keys())
        for col, header in enumerate(metrics_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        metrics = ['연환산 수익률', '연환산 변동성', '샤프 비율', '최대 낙폭 (MDD)']
        for row, metric in enumerate(metrics, 2):
            ws2.cell(row=row, column=1, value=metric).border = border
        
        for col, (name, weights) in enumerate(portfolios.items(), 2):
            if weights is None:
                continue
            
            port_returns = returns.dot(weights)
            cum_ret = (1 + port_returns).cumprod()
            total_ret = cum_ret.iloc[-1] - 1
            n_days = len(returns)
            ann_ret = (1 + total_ret) ** (252 / n_days) - 1
            ann_vol = port_returns.std() * np.sqrt(252)
            sharpe = ann_ret / ann_vol if ann_vol > 0 else 0
            mdd = ((cum_ret - cum_ret.cummax()) / cum_ret.cummax()).min()
            
            values = [ann_ret, ann_vol, sharpe, mdd]
            formats = [pct_format, pct_format, num_format, pct_format]
            
            for row, (val, fmt) in enumerate(zip(values, formats), 2):
                cell = ws2.cell(row=row, column=col, value=val)
                cell.number_format = fmt
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
        
        ws2.column_dimensions['A'].width = 18
        for col in range(2, len(portfolios) + 2):
            ws2.column_dimensions[get_column_letter(col)].width = 15
        
        # ========== Sheet 3: 데이터 정보 ==========
        ws3 = wb.create_sheet("데이터 정보")
        
        info_data = [
            ['항목', '값'],
            ['분석 일시', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['데이터 형식', results['detected_format']],
            ['시작일', price_data.index[0].strftime('%Y-%m-%d')],
            ['종료일', price_data.index[-1].strftime('%Y-%m-%d')],
            ['거래일 수', len(returns)],
            ['종목 수', len(tickers)],
            ['종목 목록', ', '.join(tickers)],
        ]
        
        for row, (key, val) in enumerate(info_data, 1):
            cell1 = ws3.cell(row=row, column=1, value=key)
            cell2 = ws3.cell(row=row, column=2, value=val)
            cell1.border = border
            cell2.border = border
            if row == 1:
                cell1.font = header_font
                cell1.fill = header_fill
                cell2.font = header_font
                cell2.fill = header_fill
        
        ws3.column_dimensions['A'].width = 15
        ws3.column_dimensions['B'].width = 40
        
        # ========== Sheet 4: 상관관계 ==========
        ws4 = wb.create_sheet("상관관계")
        
        corr_matrix = returns.corr()
        
        # 헤더
        ws4.cell(row=1, column=1, value='').border = border
        for col, ticker in enumerate(tickers, 2):
            cell = ws4.cell(row=1, column=col, value=ticker)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        for row, ticker in enumerate(tickers, 2):
            cell = ws4.cell(row=row, column=1, value=ticker)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # 상관계수 값
        for i, ticker1 in enumerate(tickers):
            for j, ticker2 in enumerate(tickers):
                cell = ws4.cell(row=i+2, column=j+2, value=corr_matrix.loc[ticker1, ticker2])
                cell.number_format = '0.00'
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                
                # 색상 표시 (높은 상관관계)
                if i != j and abs(corr_matrix.loc[ticker1, ticker2]) > 0.7:
                    cell.fill = PatternFill('solid', fgColor='FFEB9C')
        
        ws4.column_dimensions['A'].width = 10
        for col in range(2, len(tickers) + 2):
            ws4.column_dimensions[get_column_letter(col)].width = 10
        
        wb.save(file_path)

    def update_plots(self, results):
        self.figure.clear()
        ((ax1, ax2), (ax3, ax4)) = self.figure.subplots(2, 2)
        
        portfolios = results['portfolios']
        returns = results['returns']
        cov_matrix = results['cov_matrix']
        tickers = results['tickers']
        
        valid_portfolios = {k: v for k, v in portfolios.items() if v is not None}
        if not valid_portfolios:
            return
        
        colors = plt.cm.Set2(np.linspace(0, 1, len(valid_portfolios)))
        
        # 1. 가중치 비교
        x = np.arange(len(tickers))
        width = 0.8 / len(valid_portfolios)
        offset = -width * (len(valid_portfolios) - 1) / 2
        
        for (name, weights), color in zip(valid_portfolios.items(), colors):
            ax1.bar(x + offset, weights, width, label=name, color=color)
            offset += width
        
        ax1.set_xticks(x)
        ax1.set_xticklabels(tickers, rotation=45, ha='right')
        ax1.set_title('Portfolio Weights', fontweight='bold')
        ax1.legend(fontsize=8)
        ax1.set_ylabel('Weight')
        ax1.grid(axis='y', alpha=0.3)
        
        # 2. 누적 수익률
        for (name, weights), color in zip(valid_portfolios.items(), colors):
            cum_ret = (1 + returns.dot(weights)).cumprod()
            ax2.plot(cum_ret.index, cum_ret.values, label=name, color=color, linewidth=1.5)
        ax2.set_title('Cumulative Returns', fontweight='bold')
        ax2.legend(fontsize=8)
        ax2.tick_params(axis='x', rotation=45)
        ax2.grid(alpha=0.3)
        
        # 3. 리스크 기여도
        offset = -width * (len(valid_portfolios) - 1) / 2
        cov_array = cov_matrix.values
        for (name, weights), color in zip(valid_portfolios.items(), colors):
            port_vol = np.sqrt(np.dot(weights.T, np.dot(cov_array, weights)))
            if port_vol > 0:
                risk_contrib = np.multiply(np.dot(cov_array, weights), weights) / port_vol
            else:
                risk_contrib = np.zeros(len(weights))
            ax3.bar(x + offset, risk_contrib, width, label=name, color=color)
            offset += width
        
        ax3.set_xticks(x)
        ax3.set_xticklabels(tickers, rotation=45, ha='right')
        ax3.set_title('Risk Contributions', fontweight='bold')
        ax3.legend(fontsize=8)
        ax3.grid(axis='y', alpha=0.3)
        
        # 4. Risk-Return
        for (name, weights), color in zip(valid_portfolios.items(), colors):
            cum_ret = (1 + returns.dot(weights)).cumprod()
            total_ret = cum_ret.iloc[-1] - 1
            ann_ret = (1 + total_ret) ** (252 / len(returns)) - 1
            ann_vol = returns.dot(weights).std() * np.sqrt(252)
            ax4.scatter(ann_vol, ann_ret, label=name, color=color, s=150, edgecolors='black')
            ax4.annotate(name, (ann_vol, ann_ret), textcoords="offset points", xytext=(8, 5), fontsize=9)
        
        ax4.set_xlabel('Annual Volatility')
        ax4.set_ylabel('Annual Return')
        ax4.set_title('Risk-Return Profile', fontweight='bold')
        ax4.grid(True, alpha=0.3)
        
        self.figure.tight_layout()
        self.canvas.draw()


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = PortfolioOptimizer()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
